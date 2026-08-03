"""Find the molecules common to two lists and write them to a new sheet.

Given two Excel files - one "potential molecules" list and one "high priority +
monopoly" list - this extracts the molecules present in BOTH (matched by the
'Molecule' name, case-insensitive) and writes a workbook with three sheets:
Potential, High Priority Monopoly, and Common Molecules.

The extraction is defensive: each file may be a real .xlsx, a multi-sheet
workbook, or a Molecule Explorer ".xls" that is actually HTML, and the molecule
table may sit under a title row.
"""

import os
import re
import time
from pathlib import Path

import pandas as pd

MOLECULE_COL = "Molecule"
_LOG_PATH = os.path.join(os.path.expanduser("~"), "CommonMolecules_debug.log")


def _log(message: str) -> None:
    """Append a timestamped line to a debug log (flushed), so a hang is visible."""
    try:
        with open(_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"{time.strftime('%H:%M:%S')}  {message}\n")
            f.flush()
    except Exception:
        pass


def _safe(value):
    """Convert a cell value to something xlsxwriter accepts (NaN/None -> '')."""
    if value is None or (isinstance(value, float) and value != value):  # NaN check
        return ""
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return value
    return value


def _read_sheet_fast(path, sheet_name=None, max_empty_run=40, max_rows=1_000_000, max_cols=256):
    """Read one worksheet into a header-less DataFrame, fast and hang-proof.

    Bounds BOTH dimensions: reads at most `max_cols` columns (so a phantom column
    range up to ~16,384 wide never builds a giant table) and stops after a run of
    empty rows (so a phantom row range up to ~1,048,576 doesn't make it crawl).
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows, empty_run = [], 0
        for row in ws.iter_rows(values_only=True, max_col=max_cols):
            if row is None or all(_is_empty_cell(c) for c in row):
                empty_run += 1
                if empty_run >= max_empty_run:
                    break
                continue
            empty_run = 0
            rows.append(list(row))
            if len(rows) >= max_rows:
                break
    finally:
        wb.close()
    width = max((len(r) for r in rows), default=0)
    rows = [r + [None] * (width - len(r)) for r in rows]
    df = pd.DataFrame(rows)
    _log(f"_read_sheet_fast -> {df.shape[0]} rows x {df.shape[1]} cols (capped at {max_cols})")
    return df


def _named_raw_frames(path):
    """Return [(sheet_name, raw_dataframe_without_header), ...] for any file type."""
    path = str(path)
    try:
        frames = pd.read_html(path, header=None)  # HTML-disguised-as-.xls exports
        return [(f"table{i}", df) for i, df in enumerate(frames)]
    except (ValueError, ImportError):
        xls = pd.ExcelFile(path)
        return [(s, _read_sheet_fast(path, s)) for s in xls.sheet_names]


def _is_empty_cell(value):
    return value is None or (isinstance(value, str) and not value.strip()) or (
        isinstance(value, float) and value != value  # NaN
    )


def _build_table(raw: pd.DataFrame, header_row: int) -> pd.DataFrame:
    row_vals = raw.iloc[header_row].tolist()
    keep_idx = [j for j, c in enumerate(row_vals) if not _is_empty_cell(c)]
    header = [str(row_vals[j]).strip() for j in keep_idx]
    body = raw.iloc[header_row + 1:, keep_idx].copy()
    body.columns = header
    return body.dropna(how="all").reset_index(drop=True)


def _find_table(raw: pd.DataFrame):
    """Find the header row and return the table below it.

    Prefers a row containing 'Molecule'; otherwise the first row with 2+ real
    cells (skips one-cell title rows); otherwise the first non-empty row.
    """
    limit = min(len(raw), 15)
    # 1) a row that has a 'Molecule' header
    for i in range(limit):
        vals = raw.iloc[i].tolist()
        if any(not _is_empty_cell(c) and str(c).strip().lower() == "molecule" for c in vals):
            return _build_table(raw, i)
    # 2) first row with 2+ real cells (skip single-cell titles)
    for i in range(limit):
        vals = raw.iloc[i].tolist()
        if sum(not _is_empty_cell(c) for c in vals) >= 2:
            return _build_table(raw, i)
    # 3) first non-empty row
    for i in range(limit):
        vals = raw.iloc[i].tolist()
        if any(not _is_empty_cell(c) for c in vals):
            return _build_table(raw, i)
    return None


def _name_column(df: pd.DataFrame):
    """The column holding the molecule/drug name: a 'Molecule' column, else column A."""
    for col in df.columns:
        if str(col).strip().lower() == "molecule":
            return col
    return df.columns[0]


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip().lower()


def _base(value) -> str:
    """Smart key: drop strength/form (everything from ' - ' or the first digit)."""
    return re.split(r"\s*[-–]\s*|\s+\d", _norm(value))[0].strip()


def list_sheets(path):
    """Return the sheet names of a real spreadsheet, or [] for an HTML-.xls export."""
    try:
        return pd.ExcelFile(path).sheet_names
    except Exception:
        return []  # HTML-disguised-as-.xls has a single table, nothing to choose


def extract_molecule_table(path, prefer_keywords=(), sheet_name=None) -> pd.DataFrame:
    """Return a DataFrame with a 'Molecule' column.

    If `sheet_name` is given, read exactly that sheet. Otherwise auto-detect,
    preferring sheets whose name matches one of `prefer_keywords`.
    """
    if sheet_name:
        _log(f"reading sheet '{sheet_name}' of {Path(path).name} (fast/read-only)…")
        raw = _read_sheet_fast(path, sheet_name)
        _log(f"read {len(raw)} rows from '{sheet_name}'")
        table = _find_table(raw)
        if table is not None and len(table.columns):
            _log(f"table: {table.shape[0]} rows, name column = {_name_column(table)!r}")
            return table
        raise ValueError(f"No data table found in sheet '{sheet_name}' of {Path(path).name}")

    frames = _named_raw_frames(path)

    def preferred_first(named_frame):
        name = named_frame[0].lower()
        return 0 if any(k in name for k in prefer_keywords) else 1

    for _, raw in sorted(frames, key=preferred_first):
        table = _find_table(raw)
        if table is not None and len(table.columns):
            return table
    raise ValueError(f"No data table found in {Path(path).name}")


def find_common(potential_df: pd.DataFrame, hp_df: pd.DataFrame, mode: str = "smart") -> pd.DataFrame:
    """Rows of hp_df whose molecule name also appears in the potential list.

    mode:
      "smart"    - strip strength/form from the potential names, then match (default)
      "exact"    - match on the cleaned full text
      "contains" - a high-priority molecule name appears inside a potential entry
    """
    pcol = _name_column(potential_df)
    hcol = _name_column(hp_df)
    _log(f"matching mode={mode}: potential col={pcol!r}, hp col={hcol!r}")

    if mode == "contains":
        potential_norms = [_norm(v) for v in potential_df[pcol]]
        def is_common(h):
            hk = _norm(h)
            return bool(hk) and any(hk in p for p in potential_norms)
    else:
        key = _base if mode == "smart" else _norm
        potential_keys = {key(v) for v in potential_df[pcol]}
        potential_keys.discard("")
        def is_common(h):
            return _norm(h) in potential_keys

    in_both = hp_df[hcol].apply(is_common)
    result = hp_df[in_both].copy().reset_index(drop=True)
    _log(f"find_common -> {len(result)} molecules")
    return result


def _write_sheet(ws, book, df, title=None):
    """Write a DataFrame to an xlsxwriter worksheet (fast), optional bold title on row 0."""
    bold = book.add_format({"bold": True})
    start = 0
    if title is not None:
        ws.write(0, 0, title, book.add_format({"bold": True, "font_size": 13}))
        start = 1
    for c, col in enumerate(df.columns):
        ws.write(start, c, str(col), bold)
    for r, row in enumerate(df.itertuples(index=False), start=start + 1):
        for c, val in enumerate(row):
            ws.write(r, c, _safe(val))
    for idx in range(len(df.columns)):
        col_name = str(df.columns[idx])
        sample = df.iloc[:, idx].head(200).tolist()
        width = max([len(col_name)] + [len(str(v)) for v in sample])
        ws.set_column(idx, idx, min(48, width + 2))


def write_comparison_workbook(out_path, potential_df, hp_df, common_df):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # xlsxwriter is much faster than openpyxl for writing large sheets.
    import xlsxwriter
    book = xlsxwriter.Workbook(str(out_path), {"constant_memory": True})
    _write_sheet(book.add_worksheet("Potential"), book, potential_df)
    _log(f"wrote Potential ({len(potential_df)} rows)")
    _write_sheet(book.add_worksheet("High Priority Monopoly"), book, hp_df)
    _log(f"wrote High Priority Monopoly ({len(hp_df)} rows)")
    _write_sheet(book.add_worksheet("Common Molecules"), book, common_df,
                 title=f"Molecules in BOTH lists: {len(common_df)}")
    _log(f"wrote Common Molecules ({len(common_df)} rows)")
    book.close()
    return out_path


def compare_files(potential_path, hp_path, out_path, potential_sheet=None, hp_sheet=None, mode="smart"):
    """High-level: read both files, intersect, write the workbook. Returns (out_path, count).

    `potential_sheet` / `hp_sheet` optionally pin which sheet to read from each file.
    `mode` is the matching rule: "smart" (default), "exact", or "contains".
    """
    _log(f"=== compare start (mode={mode}) ===")
    _log(f"potential={os.path.basename(str(potential_path))} sheet={potential_sheet!r}")
    _log(f"hp={os.path.basename(str(hp_path))} sheet={hp_sheet!r}")
    potential_df = extract_molecule_table(potential_path, prefer_keywords=("potential",), sheet_name=potential_sheet)
    _log(f"extracted potential: {len(potential_df)} rows, cols={list(potential_df.columns)[:6]}")
    hp_df = extract_molecule_table(hp_path, prefer_keywords=("monopoly", "priority"), sheet_name=hp_sheet)
    _log(f"extracted hp: {len(hp_df)} rows, cols={list(hp_df.columns)[:6]}")
    common_df = find_common(potential_df, hp_df, mode=mode)
    write_comparison_workbook(out_path, potential_df, hp_df, common_df)
    _log(f"=== compare done -> {out_path} ===")
    return out_path, len(common_df)
