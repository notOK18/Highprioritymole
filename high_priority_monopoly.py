"""Gather high-priority, monopoly-competition molecules into their own sheet.

Reads the Molecule Explorer export, selects every molecule whose Priority is
"High" and whose Competition is "Monopoly", and writes an Excel workbook with:
  * "All Molecules"            - the full source table
  * "High Priority Monopoly"   - just the selected molecules, as a titled table

Note: the Molecule Explorer ".xls" export is actually an HTML table, so it is
read with pandas.read_html (falling back to read_excel for real spreadsheets).

Usage:
    python high_priority_monopoly.py
    python high_priority_monopoly.py --file "Data/molecule_explorer (2).xls"
    python high_priority_monopoly.py --priority High --competition monopoly
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
PRIORITY_COL = "Priority"
COMPETITION_COL = "Competition"
SORT_COL = "Units_2025"  # rank the table by market size when available


def find_default_file() -> Path:
    data_dir = HERE / "Data"
    candidates = [
        p for p in list(data_dir.glob("*.xls")) + list(data_dir.glob("*.xlsx"))
        if not p.name.startswith("~$") and not p.name.startswith(".")
    ]
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"No .xls/.xlsx file found in {data_dir}")
    return candidates[0]


def load_molecules(path: Path) -> pd.DataFrame:
    """Load the molecule table, handling the HTML-disguised-as-.xls export."""
    try:
        return pd.read_html(path)[0]
    except (ValueError, ImportError):
        return pd.read_excel(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Select high-priority monopoly molecules.")
    parser.add_argument("--file", help="Source export (default: newest in Data/).")
    parser.add_argument("--priority", default="High", help="Priority to match (default: High).")
    parser.add_argument("--competition", default="monopoly", help="Competition keyword (default: monopoly).")
    parser.add_argument("--out", default=None,
                        help="Output workbook path (default: a real .xlsx next to the source, same name).")
    args = parser.parse_args()

    path = Path(args.file) if args.file else find_default_file()
    if not path.is_absolute():
        path = HERE / path

    df = load_molecules(path)
    for needed in (PRIORITY_COL, COMPETITION_COL):
        if needed not in df.columns:
            raise ValueError(f"Column '{needed}' not found. Columns are: {list(df.columns)}")

    priority_match = df[PRIORITY_COL].astype(str).str.strip().str.lower() == args.priority.lower()
    competition_match = df[COMPETITION_COL].astype(str).str.contains(args.competition, case=False, na=False)
    selected = df[priority_match & competition_match].copy()

    if SORT_COL in selected.columns:
        selected = selected.sort_values(
            SORT_COL, ascending=False, key=lambda s: pd.to_numeric(s, errors="coerce")
        )
    selected = selected.reset_index(drop=True)

    # Default: a real Excel workbook beside the source, same base name.
    # (The source .xls is HTML and cannot itself hold multiple sheets.)
    if args.out:
        out_path = Path(args.out) if Path(args.out).is_absolute() else HERE / args.out
    else:
        out_path = path.with_suffix(".xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    title = f"High-priority molecules with {args.competition.title()} competition  ({len(selected)} molecules)"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Molecules", index=False)
        selected.to_excel(writer, sheet_name="High Priority Monopoly", index=False, startrow=1)

        ws = writer.sheets["High Priority Monopoly"]
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
        for cell in ws[2]:  # header row sits on row 2 (title is row 1)
            cell.font = Font(bold=True)
        for idx, column in enumerate(selected.columns, start=1):
            width = max(len(str(column)), *(len(str(v)) for v in selected[column].astype(str))) if len(selected) else len(str(column))
            ws.column_dimensions[get_column_letter(idx)].width = min(48, width + 2)

    print(f"Source            : {path.name}")
    print(f"Total molecules   : {len(df)}")
    print(f"Priority = {args.priority!r} AND Competition ~ {args.competition!r}: {len(selected)} molecules")
    print(f"Written to        : {out_path}")
    print(f"  Sheets          : 'All Molecules', 'High Priority Monopoly'")


if __name__ == "__main__":
    main()
